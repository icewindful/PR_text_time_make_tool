import convert_subtitle # for convert subtile function
import convert_time_word_to_sec # for convert subtile function

import os
import Globalvar as GVar # python is not "extern method" , use Global Module use it.
import window_GUI_setting as GUI
import re

from openpyxl import Workbook

import tkinter as tk
from tkinter import filedialog
# from tkinter import *
from tkinter import scrolledtext

#only python add globalvar file and all combine other file
GVar._init()

def select_file_name_is_defined():
    try:
        select_file_name
    except NameError:
        return False
    else:
        return True

def select_file():
    # select file path 
    global select_file_name 
    #define last select covert file Flag
    
    select_file_name = filedialog.askopenfilename(title='select_path')
    entry.set(select_file_name)
    print(select_file_name) # show filepath in console 


def convert_subtitle_file():
    # print(select_file_name)
    if select_file_name_is_defined() == True:
        convert_subtitle.convet_txt_type_function(select_file_name)
        GVar.select_file_state = "EDIT"
    else:
        print("this file type cann't read it")

def convert_time_file():
    # print(select_file_name)
    if select_file_name_is_defined() == True:
        convert_time_word_to_sec.process_file(select_file_name)
        GVar.select_file_state = "TIME"
    else:
        print("this file type cann't read it")


def show_Last_Edit_file():

    if select_file_name_is_defined() == True:
        split_tup = os.path.splitext(select_file_name)
        # print(select_file_name)
        print(GVar.select_file_state)
        # first_file_name = split_tup[0]
        file_type = split_tup[1]

        if file_type == ".txt" :
            if GVar.select_file_state == "unselect":
                print("unselect convert file")
                last_edit_file_name = select_file_name
                pass
            if GVar.select_file_state == "EDIT":
                print("last file state = EDIT")
                last_edit_file_name = select_file_name.replace('.txt', 'EDIT.txt')
            if GVar.select_file_state == "TIME":
                print("last file state= TIME")
                last_edit_file_name = select_file_name.replace('.txt', '_TIME.txt')
            
            show_Edit_file = open(last_edit_file_name,'r', encoding='utf-8')
            s = show_Edit_file.read()
            # print(s)
            show_subtitle_txt.delete("1.0","end")
            show_subtitle_txt.insert("end",s)
        else:
            print("this file type cann't read it")
    else:
        print("this file type cann't read it")

def show_Select_file():
    if select_file_name_is_defined() == True:
        split_tup = os.path.splitext(select_file_name)
        # print(select_file_name)

        # first_file_name = split_tup[0]
        file_type = split_tup[1]

        if file_type == ".txt" :
            show_Edit_file = open(select_file_name,'r', encoding='utf-8')
            s = show_Edit_file.read()
            # print(s)
            show_subtitle_txt.delete("1.0","end")
            show_subtitle_txt.insert("end",s)
        else:
            print("this file type cann't read it")
    else:
        print("this file type cann't read it")

def show_Edit_clear():
    show_subtitle_txt.delete("1.0","end")

def test_funtion():
    print("test")
    wb = Workbook()
    ws = wb.active

    #ws['A1'] = 42
    ws.cell(row=1,column=1,value="test")


    ws.append([1,2,3])
    import datetime
    ws['A3'] = datetime.datetime.now()

    wb.save("sample.xlsx")
    wb.close()

def time_to_excel():
    if select_file_name_is_defined() == True:
        wb = Workbook()
        ws = wb.active # open work sheet
        r = 2 # row
        c = 2 # column
        shift = 1
        
        list_text_shift = c 

        split_text_list = split_entry.get().split(",")
        split_text_list_clear_ip = [ clear_space.strip() for clear_space in split_text_list if clear_space.strip()!='']

        print(split_text_list_clear_ip)

        set_text = "Icewindful"
        try:
            print(split_text_list_clear_ip.index(set_text))
        except ValueError:
            print("Value_Type")
        except :
            pass

        with open(select_file_name, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            for line in lines:
                # print(line)
                if line in ['\n','\r\n']: # remove space
                    pass
                else :
                    time = re.findall(r'[〈](.*?)[〉]',line)
                    ws.cell(row=r,column=c,value=time[0])
                    splittext = line.split('【')
                    infotext = "【" + splittext[-1]

                    class_text = re.findall(r'[【](.*?)[】]',infotext)
                    class_ip = [ clear_space.strip() for clear_space in class_text if clear_space.strip()!='']
                    # print(class_ip)

                    try:
                        print(split_text_list_clear_ip.index(class_ip[0]))
                        list_text_shift = shift + split_text_list_clear_ip.index(class_ip[0]) + 1
                    except ValueError:
                        list_text_shift = shift
                        # print("not find Value")
                    except :
                        pass
                    
                    # print(infotext)

                    # ws.cell(row=r,column=(c+shift),value=infotext)
                    ws.cell(row=r,column=(c+list_text_shift),value=infotext)
                    # https://youtu.be/JLwrqscIlGM&t=
                    # ws.cell(row=r,column=(c+shift),value=infotext).hyperlink = "https://youtu.be/JLwrqscIlGM&t="+time[0]
                    ws.cell(row=r,column=(c+list_text_shift),value=infotext).hyperlink = "https://youtu.be/JLwrqscIlGM&t="+time[0]
                    #print(infotext)
                    #print(time)
                    
                    r = r + 1
            
        wb.save("sample.xlsx")
        wb.close()
    else:
        print("this file type cann't read it")


# make windows gui
window = tk.Tk()

window.title(GUI.top_title_text)
window.geometry(GUI.window_geometry)

# show select file path
file_path_label = tk.Label(window, text = GUI.File_Path_Name , fg = "Red")
file_path_label.place(x=GUI.File_Path_Name_X , y = GUI.File_Path_Name_Y )

entry = tk.StringVar()
entry.set(GUI.show_text_word)
entry_box = tk.Entry(window, textvariable=entry, width=GUI.show_text_word_width)
# ※pack and grid cann't use same block in windows ※
entry_box.place(x= GUI.show_text_word_X ,y= GUI.show_text_word_Y)

# creat select file button 
button_loadpath = tk.Button(window, text=GUI.button_loadpath_name, command=select_file , padx=2, pady=2 ,cursor = "hand2")
button_loadpath.place(x = GUI.button_loadpath_name_X , y = GUI.button_loadpath_name_Y)


# creat covnert text button 
button_convert_subtitle = tk.Button(window, text=GUI.button_convert_subtitle_name, command=convert_subtitle_file , padx=2, pady=2 ,cursor = "hand2" ,fg="blue")
button_convert_subtitle.place(x = GUI.button_convert_subtitle_name_X , y = GUI.button_convert_subtitle_name_Y)

button_show_Edit_button = tk.Button(window, text=GUI.button_show_Select_button_name, command=show_Select_file , padx=2, pady=2 ,cursor = "hand2")
button_show_Edit_button.place(x = GUI.button_show_Select_button_name_X , y = GUI.button_show_Select_button_name_Y)

button_show_Edit_button = tk.Button(window, text=GUI.button_show_Edit_button_name, command=show_Last_Edit_file , padx=2, pady=2 ,cursor = "hand2")
button_show_Edit_button.place(x = GUI.button_show_Edit_button_name_X , y = GUI.button_show_Edit_button_name_Y)

button_show_Clear_button = tk.Button(window, text=GUI.button_show_Clear_button_name, command=show_Edit_clear , padx=2, pady=2 ,cursor = "hand2")
button_show_Clear_button.place(x = GUI.button_show_Clear_button_name_X , y = GUI.button_show_Clear_button_name_Y)

button_time_button = tk.Button(window, text=GUI.button_time_convert_sec_name, command=convert_time_file , padx=2, pady=2 ,cursor = "hand2")
button_time_button.place(x = GUI.button_time_convert_sec_name_X , y = GUI.button_time_convert_sec_name_Y)


### split title load text
split_text_list_label = tk.Label(window, text = GUI.split_text_list_label_name , fg = "Red")
split_text_list_label.place(x=GUI.split_text_list_label_X , y = GUI.split_text_list_label_Y )

split_entry = tk.StringVar()
split_entry.set(GUI.enter_split_list_box_sample_text)
split_entry_box = tk.Entry(window, textvariable=split_entry, width=GUI.enter_split_list_box_width)
# ※pack and grid cann't use same block in windows ※
split_entry_box.place(x= GUI.enter_split_list_box_X ,y= GUI.enter_split_list_box_Y)

#for test button
write_to_excel = tk.Button(window, text=GUI.test_function_name, command=time_to_excel , padx=2, pady=2 ,cursor = "hand2")
write_to_excel.place(x = GUI.test_function_name_X , y = GUI.test_function_name_Y)

#below show text
show_subtitle_txt = scrolledtext.ScrolledText(window,width=GUI.scroll_text_box_width,height=GUI.scroll_text_box_height)
show_subtitle_txt.place(x = GUI.scroll_text_box_X , y = GUI.scroll_text_box_Y)


# keep windows 
window.mainloop()